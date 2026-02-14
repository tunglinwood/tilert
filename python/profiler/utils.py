import os
from dataclasses import dataclass
from typing import Any

import numpy as np
import torch

from tilert.utils import SLICES_FOR_TILERT_OP

# Worker names used by ExecPlanDescriptor (previously from scheduling.plan_v0)
WORKER_NAMES = [
    "Init",
    "Prefetch",
    "Compute",
    "ExtraTask1/SyncIo",
    "ExtraTask2/IoP0",
    "ExtraTask3/IoP2",
    "ExtraTask4",
    "ExtraTask5",
]

try:
    from openpyxl import Workbook
    from openpyxl.cell import Cell
    from openpyxl.styles import Alignment, Border, PatternFill, Side
    from openpyxl.styles.colors import COLOR_INDEX
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("openpyxl is not installed, profile logs will not be visualized")
    Workbook = None


__all__ = [
    "ExcelStyleConfigs",
    "ExecPlanDescriptor",
    "WorkerBookVisualizer",
    "visualize_profile_logs",
    "parse_profile_log_tensor",
    "parse_op_time",
]


@dataclass
class ExcelStyleConfigs:
    """Excel style configurations."""

    # 2 col * 3 stream
    cols_per_worker: int = 6
    ns_per_tick: int = 1000


@dataclass
class ExecPlanDescriptor:
    """Exec plan descriptor."""

    workers_def: list
    op_lists: list


class WorkerBookVisualizer:
    """Sheet visualizer."""

    def __init__(self, exec_plan_desc: ExecPlanDescriptor):
        self.exec_plan_desc = exec_plan_desc

        self.wb = Workbook()
        self.wb.remove(self.wb.active)

        # Excel configs
        self.style_configs = ExcelStyleConfigs()

        self.op_cols_splits = 3

        self.time_bar_cols = 1
        self.op_stat_bar_cols = 6

        workers_num = len(self.exec_plan_desc.workers_def)
        self.op_vis_bar_cols = workers_num * self.style_configs.cols_per_worker
        assert self.op_stat_bar_cols % self.op_cols_splits == 0

    @property
    def time_bar_next_col(self) -> int:
        return self.time_bar_cols + 1

    @property
    def op_stat_bar_next_col(self) -> int:
        return self.time_bar_next_col + self.op_stat_bar_cols

    @property
    def op_vis_bar_next_col(self) -> int:
        return self.op_stat_bar_next_col + self.op_vis_bar_cols

    @staticmethod
    def add_region_cell(
        ws: Worksheet,
        value: str,
        start_row: int,
        start_col: int,
        row_size: int = 1,
        col_size: int = 1,
        color_offset: int = -1,
    ) -> Cell:
        cell = ws.cell(row=start_row, column=start_col, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if color_offset >= 0:
            cell.fill = PatternFill(
                start_color=COLOR_INDEX[50 + color_offset],
                end_color=COLOR_INDEX[50 + color_offset],
                fill_type="solid",
            )
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=start_row + row_size - 1,
            end_column=start_col + col_size - 1,
        )
        return cell

    def init_layout(self, ws: Worksheet) -> None:
        workers_name = self.exec_plan_desc.workers_def
        worker_cols = self.style_configs.cols_per_worker

        self.add_region_cell(ws, "Op Info", 1, self.time_bar_next_col, 1, self.op_stat_bar_cols)

        for worker_id, worker_name in enumerate(workers_name):
            start_col = worker_cols * worker_id + self.op_stat_bar_next_col
            self.add_region_cell(ws, worker_name, 1, start_col, 1, worker_cols)

    def _parse_inst_info(
        self, insts_info: list[tuple[str, float, int] | tuple[str, float] | str], op_idx: int
    ) -> tuple[str, float, int]:
        inst_info = insts_info[op_idx]
        if isinstance(inst_info, str):
            op_name, op_cost = inst_info, 0.0
            op_stream = op_idx % self.op_cols_splits
        elif len(inst_info) == 2:
            op_name, op_cost = inst_info
            op_stream = op_idx % self.op_cols_splits
        elif len(inst_info) == 3:
            op_name, op_cost, op_stream = inst_info
        else:
            raise TypeError("Invalid inst_info format")
        return op_name, op_cost, op_stream

    def add_region_cell_by_time(
        self,
        ws: Worksheet,
        op_show_info: str,
        start_time: float,
        end_time: float,
        op_col_start: int,
        op_col_size: int,
        ns_tick: int,
        color_offset: int = -1,
    ) -> Cell:
        op_start_row_idx = np.round(start_time / ns_tick).astype(np.int32) + 2
        op_end_row_idx = np.round(end_time / ns_tick).astype(np.int32) + 2
        op_end_row_idx = max(op_end_row_idx, op_start_row_idx)
        return self.add_region_cell(
            ws,
            op_show_info,
            op_start_row_idx,
            op_col_start,
            max(op_end_row_idx - op_start_row_idx, 1),
            op_col_size,
            color_offset,
        )

    def timeline_visual_region(
        self,
        ws: Worksheet,
        profile_logs: np.ndarray,
        insts_info: list[tuple[str, float, int] | tuple[str, float] | str],
        ignore_prefilling: bool = True,
    ) -> None:
        ns_tick = self.style_configs.ns_per_tick
        self.init_layout(ws)

        total_end_time = 0
        for op_idx, op_log in enumerate(profile_logs):
            op_name, op_cost, op_stream = self._parse_inst_info(insts_info, op_idx)

            if op_stream >= self.op_cols_splits:
                print(f"stream_id (aka col_id) must < {self.op_cols_splits}")
                raise ValueError

            valid_mask: np.ndarray = op_log >= 0
            if ignore_prefilling:
                valid_mask[2:4] = False

            if np.count_nonzero(valid_mask) == 0:
                continue

            op_start_time = np.min(op_log, where=valid_mask, initial=np.inf)
            op_end_time = np.max(op_log, where=valid_mask, initial=-np.inf)
            total_end_time = max(total_end_time, op_end_time)

            op_cost_theory = op_cost / 1000
            op_cost_actual = (op_end_time - op_start_time) / 1000
            op_bw_utils = f"{op_cost_theory / op_cost_actual * 100:.2f}"

            op_show_info = (
                f"{op_name}\n"
                + f"BW Util: {op_bw_utils}%\n"
                + f"Actual: {op_cost_actual:.2f}us\n"
                + f"Theoretical: {op_cost_theory:.2f}us\n"
                + f"Start Time: {op_start_time / 1000:.2f}us\n"
                + f"End Time: {op_end_time / 1000:.2f}us"
            )
            op_col_size = self.op_stat_bar_cols // self.op_cols_splits
            op_col_start = self.time_bar_next_col + op_stream * op_col_size
            self.add_region_cell_by_time(
                ws,
                op_show_info,
                op_start_time,
                op_end_time,
                op_col_start,
                op_col_size,
                ns_tick,
            )

            for queue_idx, (start_time, end_time) in enumerate(zip(op_log[::2], op_log[1::2])):
                if start_time < 0 or end_time < 0:
                    continue
                task_dur = (end_time - start_time) / 1000
                task_bw_utils = f"{min(100, op_cost_theory / task_dur * 100):.2f}"
                task_show_info = (
                    f"{op_name}\n"
                    + f"Dur: {task_dur:.2f}us\n"
                    + f"BW Util. {task_bw_utils}%:\n"
                    + f"Start: {start_time / 1000:.2f}us\n"
                    + f"End: {end_time / 1000:.2f}us"
                )
                task_col_size = self.style_configs.cols_per_worker // self.op_cols_splits
                task_col_start = (
                    self.op_stat_bar_next_col
                    + queue_idx * self.style_configs.cols_per_worker
                    + op_stream * task_col_size
                )
                cell = self.add_region_cell_by_time(
                    ws,
                    task_show_info,
                    start_time,
                    end_time,
                    task_col_start,
                    task_col_size,
                    ns_tick,
                    queue_idx,
                )
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        for dur_idx, dur_start in enumerate(range(0, int(total_end_time), ns_tick)):
            ws.cell(row=dur_idx + 2, column=1, value=f"{(dur_start + ns_tick) / 1000:.2f}")

    def brief_table_region(
        self,
        ws: Worksheet,
        profile_logs: np.ndarray,
        insts_info: list[tuple[str, float, int] | tuple[str, float] | str],
    ) -> None:
        for op_idx, op_log in enumerate(profile_logs):
            op_name, _, _ = self._parse_inst_info(insts_info, op_idx)

            ws.cell(row=op_idx + 2, column=self.op_vis_bar_next_col, value=op_name)

            for queue_idx, (start_time, end_time) in enumerate(zip(op_log[::2], op_log[1::2])):
                if start_time < 0 or end_time < 0:
                    continue
                task_dur = (end_time - start_time) / 1000
                ws.cell(
                    row=op_idx + 2, column=self.op_vis_bar_next_col + queue_idx + 1, value=task_dur
                )

    def add_sheet(self, profile_logs: np.ndarray, sheet_name: str) -> "WorkerBookVisualizer":
        """Add a sheet to the workbook."""
        wb = self.wb
        insts_info = self.exec_plan_desc.op_lists

        ws = wb.create_sheet(sheet_name)
        self.timeline_visual_region(ws, profile_logs, insts_info)
        self.brief_table_region(ws, profile_logs, insts_info)

        return self

    def add_sm_brief_sheet(
        self, profile_logs: np.ndarray, sheet_name: str
    ) -> "WorkerBookVisualizer":
        """Add a brief sheet to workbook which contains min/max start/end and duration among SMs"""
        wb = self.wb
        insts_info = self.exec_plan_desc.op_lists
        ws = wb.create_sheet(sheet_name)

        profile_logs = np.transpose(profile_logs, (1, 0, 2))

        # 1. init layout
        workers_name = self.exec_plan_desc.workers_def
        worker_metric_def = [
            "min_start",
            "max_end",
            "min_dur",
            "max_dur",
            "mean_dur",
            "std_dur",
        ]

        worker_cols = len(worker_metric_def)

        self.add_region_cell(ws, "Op Info", 1, self.time_bar_next_col, 1, self.op_stat_bar_cols)

        for worker_id, worker_name in enumerate(workers_name):
            start_col = worker_cols * worker_id + self.op_stat_bar_next_col
            self.add_region_cell(ws, worker_name, 1, start_col, 1, worker_cols)
            for metric_id, metric_name in enumerate(worker_metric_def):
                start_col_metric = start_col + metric_id
                self.add_region_cell(ws, metric_name, 2, start_col_metric, 1, 1)

        # 2. calc metrics
        # profile_logs: (num_ops, num_sm, num_task*2)
        for op_idx, op_profile_log in enumerate(profile_logs):
            valid_mask = (op_profile_log >= 0) & (op_profile_log < 1e9)
            # skip if this op is fully invalid
            if not np.any(valid_mask):
                continue

            op_name, _, _ = self._parse_inst_info(insts_info, op_idx)
            self.add_region_cell(ws, op_name, op_idx + 3, self.time_bar_next_col, 1, 2)

            for queue_idx in range(op_profile_log.shape[1] // 2):
                starts = op_profile_log[:, queue_idx * 2]
                ends = op_profile_log[:, queue_idx * 2 + 1]

                valid_mask = (
                    (starts >= 0) & (starts < 1e9) & (ends >= 0) & (ends < 1e9) & (starts <= ends)
                )

                valid_starts = starts[valid_mask] / 1000
                valid_ends = ends[valid_mask] / 1000

                if len(valid_starts) == 0:
                    continue

                min_start = np.min(valid_starts)
                max_end = np.max(valid_ends)
                durations = valid_ends - valid_starts

                metrics_values = [
                    min_start,
                    max_end,
                    np.min(durations),
                    np.max(durations),
                    np.mean(durations),
                    np.std(durations),
                ]

                # row_idx start from 3, because {1: work_name, 2: metric_name}
                # col_idx start from worker::start_col
                start_row = op_idx + 3
                start_col = worker_cols * queue_idx + self.op_stat_bar_next_col
                color_offset = queue_idx

                for i, value in enumerate(metrics_values):
                    # color mean and std dev
                    cell_color = color_offset if i >= 4 else -1
                    self.add_region_cell(ws, value, start_row, start_col + i, 1, 1, cell_color)

        return self

    def save(self, out_path: str) -> None:
        """Save the workbook to a file."""
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        self.wb.save(out_path)


def visualize_profile_logs(
    all_profile_logs: np.ndarray,
    out_path: str,
    inst2opname: list[tuple[str, float, int] | tuple[str, float] | str],
    with_mean: bool = False,
    with_max: bool = False,
) -> None:
    """Visualize profile logs."""
    valid_ctas = np.argwhere(np.any(all_profile_logs != 0, axis=(1, 2)))[:, 0]
    filtered_logs = all_profile_logs[valid_ctas]
    filtered_masks = np.logical_and(filtered_logs >= 0, filtered_logs < 1e9)
    mean_profile_logs = np.mean(filtered_logs, axis=0, where=filtered_masks)
    mean_profile_logs[np.isnan(mean_profile_logs)] = -1
    if filtered_logs.size == 0:
        return
    assemble_profile_logs = np.zeros_like(filtered_logs[0])
    assemble_profile_logs[:, ::2] = np.min(
        filtered_logs[..., ::2], axis=0, where=filtered_masks[..., ::2], initial=np.inf
    )
    assemble_profile_logs[:, 1::2] = np.max(
        filtered_logs[..., 1::2], axis=0, where=filtered_masks[..., 1::2], initial=-np.inf
    )
    assemble_profile_logs[np.isinf(assemble_profile_logs)] = -1

    visualizer = WorkerBookVisualizer(ExecPlanDescriptor(WORKER_NAMES, inst2opname))
    if with_mean:
        visualizer.add_sheet(mean_profile_logs, "mean")
    if with_max:
        raise NotImplementedError("with_max is not implemented")

    visualizer.add_sm_brief_sheet(filtered_logs, "mean_sm_brief")
    for block_idx, profile_logs in enumerate(filtered_logs):
        profile_logs[profile_logs > 1e9] = -1
        visualizer.add_sheet(profile_logs, f"block_{block_idx}")
    visualizer.save(out_path)


def parse_profile_log_tensor(
    profile_logs_tensor: torch.Tensor,
    out_path: str,
    inst2opname: Any,
    with_mean: bool = False,
) -> None:
    """Parse a profile log tensor into a dictionary.

    Args:
        profile_log_tensor: The profile log tensor.
        out_path: The path to save the profile logs.
        inst2opname: The mapping from instance index to operation name.

            list[tuple[str, float, int] | tuple[str, float] | str]

    Returns:
        None.
    """
    # Remove the extra slices for storing instructions and glb bars.
    profile_logs_tensor = profile_logs_tensor[:-SLICES_FOR_TILERT_OP, :, :]

    profile_logs = profile_logs_tensor.cpu().detach().numpy()
    valid_insts_logs = np.any(profile_logs != 0, axis=(1, 2))
    profile_logs = profile_logs[valid_insts_logs]
    valid_blocks_logs = np.any(profile_logs != 0, axis=(0, 2))
    profile_logs = profile_logs[:, valid_blocks_logs, :]
    # Return if no valid blocks logs are found.
    if profile_logs.size == 0:
        print("Warning: No profile logs available.")
        return
    profile_logs = np.transpose(profile_logs, (1, 0, 2))
    ctx_start_times = profile_logs[:, 0, 0]
    profile_logs = profile_logs[:, 1:, :]
    profile_logs = (profile_logs - ctx_start_times[:, None, None]).astype(np.float32) / 1.855

    if Workbook is not None:
        visualize_profile_logs(profile_logs, out_path, inst2opname, with_mean)


def parse_op_time(profile_logs: torch.Tensor, op_idx: int = 0, block_idx: int = 0) -> None:
    data = profile_logs[op_idx, block_idx, :].cpu().numpy()
    max_time = data.max()
    start_time = data.min()
    FREQUENCY = 1850.0

    worker_names = [
        "controller",
        "   sync_io",
        "     io_p0",
        "     io_p1",
        "     io_p2",
        "  consumer",
        "    extra1",
        "    extra2",
    ]
    for i, worker_name in enumerate(worker_names):
        if data[i * 2] != max_time:
            print(
                f"{worker_name}:\tstart:{(data[i * 2] - start_time) / FREQUENCY:.3f}, "
                f"duration:{(data[i * 2 + 1] - data[i * 2]) / FREQUENCY:.3f}, "
                f"end:{(data[i * 2 + 1] - start_time) / FREQUENCY:.3f}"
            )
